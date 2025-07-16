from pathlib import Path
import os

import yfinance as yf
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from fredapi import Fred

PROJECT_ROOT = Path(__file__).resolve().parent.parent   # adjust if needed
DATA_DIR = PROJECT_ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)

load_dotenv(PROJECT_ROOT / ".env")
FRED_KEY = os.getenv("FRED_API_KEY")
FRED = Fred(api_key=FRED_KEY)

def refresh_market_data(ticker: str):
    tkr = yf.Ticker(ticker)
    info = tkr.get_info()

    numbers = {
        "market_cap": info["marketCap"],
        "ebitda": info["ebitda"],
        "beta_raw": info["beta"],
    }

    # ------------- pull risk-free 10-yr ------------- #
    dgs10 = FRED.get_series_latest_release("DGS10").iloc[-1] / 100.0   # decimal

    # --------------- write to Excel --------------- #
    xl_path = PROJECT_ROOT / "templates" / "DCF_Model.xlsx"
    wb = load_workbook(xl_path)
    ws = wb["Inputs"]
    ws["B1"].value = ticker
    ws["B4"].value = numbers["beta_raw"]
    ws["B6"].value = dgs10
    ws["B21"].value = numbers["market_cap"]
    ws["B22"].value = numbers["ebitda"]
    wb.save(xl_path)
    print("🔄  Market data refreshed → Inputs sheet")

if __name__ == "__main__":
    refresh_market_data("T") # replace with your ticker